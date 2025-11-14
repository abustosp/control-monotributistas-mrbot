import tkinter as tk
from tkinter.messagebox import showinfo
from lib.caller_mc import consulta_mis_comprobantes
from lib.caller_rcel import consulta_rcel, validar_respuesta_rcel
from lib.utils import descargar_archivo, extraccion_urls_minio, extraer_zip, guardar_json
from lib.formatos import Aplicar_formato_encabezado, Aplicar_formato_moneda, Autoajustar_columnas, Agregar_filtros, Alinear_columnas
from dotenv import load_dotenv
import os
import pandas as pd
from datetime import date, datetime
import numpy as np
from openpyxl import load_workbook

load_dotenv()


def _fmt_fecha(val):
    """Formatea valores de fecha al formato DD/MM/YYYY."""
    if pd.isna(val):
        return ''
    if isinstance(val, (pd.Timestamp, datetime, date)):
        return val.strftime('%d/%m/%Y')
    parsed = pd.to_datetime(val, errors='coerce')
    if pd.isna(parsed):
        return str(val)
    return parsed.strftime('%d/%m/%Y')


def procesar_descarga_mc(row, mrbot_user, mrbot_api_key, base_url, mis_comprobantes_endpoint, downloads_mc_path):
    """
    Procesa la descarga de Mis Comprobantes para un contribuyente.
    
    Args:
        row: Fila del DataFrame con los datos del contribuyente
        mrbot_user: Usuario de Mrbot
        mrbot_api_key: API key de Mrbot
        base_url: URL base de la API
        mis_comprobantes_endpoint: Endpoint de Mis Comprobantes
        downloads_mc_path: Directorio de descargas
    """
    cuit_representante = str(row['CUIT_Representante'])
    clave_representante = row['Clave_representante']
    cuit_representado = str(row['CUIT_Representado'])
    desde = _fmt_fecha(row['Desde_MC'])
    hasta = _fmt_fecha(row['Hasta_MC'])
    denominacion_mc = row['Denominacion_MC']
    descarga_MC = str(row['Descarga_MC']).lower().strip()
    descarga_MC_emitidos = str(row['Descarga_MC_emitidos']).lower().strip()
    descarga_MC_recibidos = str(row['Descarga_MC_recibidos']).lower().strip()

    if descarga_MC != 'si':
        print(f"Saltando descarga MC para CUIT {cuit_representado} - Descarga_MC: {descarga_MC}")
        return
    
    print(f"\n{'='*80}")
    print(f"Procesando MC: {denominacion_mc} - CUIT: {cuit_representado}")
    print(f"Desde {desde} hasta {hasta}")
    print(f"Emitidos: {descarga_MC_emitidos} | Recibidos: {descarga_MC_recibidos}")
    print(f"{'='*80}\n")

    # Determinar qué descargar
    descargar_emitidos = descarga_MC_emitidos == 'si'
    descargar_recibidos = descarga_MC_recibidos == 'si'

    if not descargar_emitidos and not descargar_recibidos:
        print(f"No hay nada que descargar para {cuit_representado}")
        return

    try:
        # Consultar API
        response = consulta_mis_comprobantes(
            mrbot_user=mrbot_user,
            mrbot_api_key=mrbot_api_key,
            base_url=base_url,
            mis_comprobantes_endpoint=mis_comprobantes_endpoint,
            desde=desde,
            hasta=hasta,
            cuit_inicio_sesion=cuit_representante,
            representado_nombre=denominacion_mc,
            representado_cuit=cuit_representado,
            contrasena=clave_representante,
            descarga_emitidos=descargar_emitidos,
            descarga_recibidos=descargar_recibidos,
        )

        # Extraer URLs de MinIO
        urls = extraccion_urls_minio(response)

        # Crear directorio del contribuyente
        directorio_contribuyente = os.path.join(downloads_mc_path, f"{cuit_representado}_{denominacion_mc}")
        directorio_extraido = os.path.join(directorio_contribuyente, "extraido")
        os.makedirs(directorio_contribuyente, exist_ok=True)
        os.makedirs(directorio_extraido, exist_ok=True)

        # Descargar archivos
        archivos_descargados = []
        
        if descargar_emitidos and urls.get('emitidos'):
            print(f"\nDescargando emitidos...")
            archivo_emitidos = descargar_archivo(
                urls['emitidos'],
                None,
                directorio_contribuyente
            )
            archivos_descargados.append(archivo_emitidos)
        
        if descargar_recibidos and urls.get('recibidos'):
            print(f"\nDescargando recibidos...")
            archivo_recibidos = descargar_archivo(
                urls['recibidos'],
                None,
                directorio_contribuyente
            )
            archivos_descargados.append(archivo_recibidos)

        # Extraer ZIPs
        for archivo_zip in archivos_descargados:
            if archivo_zip and archivo_zip.endswith('.zip'):
                print(f"\nExtrayendo: {archivo_zip}")
                try:
                    extraer_zip(archivo_zip, directorio_extraido)
                except Exception as e:
                    print(f"Error al extraer {archivo_zip}: {e}")

        print(f"\n✓ Proceso MC completado para {denominacion_mc}")

    except Exception as e:
        print(f"\n✗ Error procesando MC {denominacion_mc} (CUIT: {cuit_representado}): {e}")


def procesar_descarga_rcel(row, mrbot_user, mrbot_api_key, base_url, rcel_endpoint, downloads_rcel_path):
    """
    Procesa la descarga de RCEL para un contribuyente.
    
    Args:
        row: Fila del DataFrame con los datos del contribuyente
        mrbot_user: Usuario de Mrbot
        mrbot_api_key: API key de Mrbot
        base_url: URL base de la API
        rcel_endpoint: Endpoint de RCEL
        downloads_rcel_path: Directorio de descargas
    """
    cuit_representante = str(row['CUIT_Representante'])
    clave_representante = row['Clave_representante']
    cuit_representado = str(row['CUIT_Representado'])
    desde = _fmt_fecha(row['Desde_RCEL'])
    hasta = _fmt_fecha(row['Hasta_RCEL'])
    denominacion_rcel = row['Denominacion_RCEL']
    descarga_RCEL = str(row['Descarga_RCEL']).lower().strip()

    if descarga_RCEL != 'si':
        print(f"Saltando descarga RCEL para CUIT {cuit_representado} - Descarga_RCEL: {descarga_RCEL}")
        return
    
    print(f"\n{'='*80}")
    print(f"Procesando RCEL: {denominacion_rcel} - CUIT: {cuit_representado}")
    print(f"Desde {desde} hasta {hasta}")
    print(f"{'='*80}\n")

    try:
        # Consultar API
        response = consulta_rcel(
            mrbot_user=mrbot_user,
            mrbot_api_key=mrbot_api_key,
            base_url=base_url,
            rcel_endpoint=rcel_endpoint,
            desde=desde,
            hasta=hasta,
            cuit_inicio_sesion=cuit_representante,
            representado_nombre=denominacion_rcel,
            representado_cuit=cuit_representado,
            contrasena=clave_representante,
        )

        # Validar respuesta
        facturas = validar_respuesta_rcel(response)

        if not facturas:
            print(f"No se encontraron facturas RCEL para {denominacion_rcel}")
            return

        # Crear directorio del contribuyente
        directorio_contribuyente = os.path.join(downloads_rcel_path, f"{cuit_representado}_{denominacion_rcel}")
        os.makedirs(directorio_contribuyente, exist_ok=True)

        # Descargar archivos
        print(f"\nDescargando {len(facturas)} facturas...")
        for factura in facturas:
            url_pdf = factura.get("URL_MINIO")
            if url_pdf:
                try:
                    ruta_descargada = descargar_archivo(
                        url_pdf,
                        None,
                        directorio_contribuyente
                    )
                    # Guardar metadata JSON
                    guardar_json(factura, ruta_descargada)
                except Exception as e:
                    print(f"Error descargando factura {factura.get('NUMERO_FACTURA', 'N/A')}: {e}")
            else:
                print(f"Factura sin URL_MINIO: {factura.get('NUMERO_FACTURA', 'N/A')}")

        print(f"\n✓ Proceso RCEL completado para {denominacion_rcel}")

    except Exception as e:
        print(f"\n✗ Error procesando RCEL {denominacion_rcel} (CUIT: {cuit_representado}): {e}")
        

def control(
    archivos_mc: str ,
    archivos_PDF: str,
    archivos_PDF_JSON: str
    ):
    '''
    Controla los datos de los archivos de 'Mis Comprobantes' con las escalas de categorías de AFIP


    '''
    # Mostrar mensaje de inicio
    #showinfo("Control de datos", "Se iniciará el control de los datos de los archivos de 'Mis Comprobantes' con las escalas de categorías de AFIP y los PDF de las facturas")

    # Leer Excel con las tablas de las escalas
    categorias = pd.read_excel('Categorias.xlsx')

    # Leer la celda 'A2' de la hoja 'Rango de Fechas' y guardarla en la variable 'fecha_inicial' en formato datetime
    fecha_inicial = pd.read_excel('Categorias.xlsx', sheet_name='Rango de Fechas', header=None, skiprows=1, usecols=[0]).iloc[0,0]
    fecha_inicial = pd.to_datetime(fecha_inicial , format='%d/%m/%Y')
    # leer la celda 'B2' en fomato fecha
    fecha_final = pd.read_excel('Categorias.xlsx', sheet_name='Rango de Fechas', header=None, skiprows=1, usecols=[1]).iloc[0,0]
    fecha_final = pd.to_datetime(fecha_final , format='%d/%m/%Y')

    # Crear un DataFrame vacio para guardar los datos consolidados
    consolidado = pd.DataFrame()
    Info_Facturas_PDF = pd.DataFrame()

    # Leer cada uno de los archivos de 'Mis Comprobantes' y concat en el DataFrame Consolidado
    # Consolidar archivos y renombrar columnas
    # consolidadar columnas
    for f in archivos_mc:
        #Si el existe el archivo, leerlo
        if os.path.isfile(f):  
            data = pd.read_csv(f, sep=';', decimal=',', encoding='utf-8-sig' )
            # si el datsaframe esta vacio, no hacer nada
            if len(data) > 0:
                # Crear la columna 'Archivo' con el ultimo elemento de 'f' separado por "/"
                data['Archivo'] = f.split("/")[-1]
                
                # Extraer información del nombre del archivo
                # Formato: NumArchivo - Tipo - FechaDesde - FechaHasta - CUIT - Denominacion.csv
                partes_archivo = data["Archivo"].str.split("-")
                data['Fin CUIT'] = partes_archivo.str[4].str.strip().astype(np.int64)  # CUIT del contribuyente
                data['CUIT Cliente'] = partes_archivo.str[4].str.strip().astype(np.int64)  # Mismo CUIT
                data['Cliente'] = partes_archivo.str[5].str.strip().str.replace('.csv','', regex=True)
                
                # Detectar si es MCE (emitidos) o MCR (recibidos) por las columnas
                es_emitido = 'Denominación Receptor' in data.columns
                es_recibido = 'Denominación Emisor' in data.columns
                
                # Unificar las columnas según el tipo
                if es_emitido:
                    # Para emitidos: Receptor es el cliente
                    data['Nro. Doc. Receptor/Emisor'] = data['Nro. Doc. Receptor']
                    data['Denominación Receptor/Emisor'] = data['Denominación Receptor']
                elif es_recibido:
                    # Para recibidos: Emisor es el proveedor
                    data['Nro. Doc. Receptor/Emisor'] = data['Nro. Doc. Emisor']
                    data['Denominación Receptor/Emisor'] = data['Denominación Emisor']
                
                # Seleccionar solo las columnas necesarias para el consolidado
                columnas_necesarias = [
                    'Fecha de Emisión', 'Tipo de Comprobante', 'Punto de Venta', 
                    'Número Desde', 'Número Hasta', 'Cód. Autorización',
                    'Tipo Cambio', 'Moneda', 
                    'Imp. Neto Gravado Total', 'Imp. Neto No Gravado', 
                    'Imp. Op. Exentas', 'Otros Tributos', 'Total IVA', 'Imp. Total',
                    'Nro. Doc. Receptor/Emisor', 'Denominación Receptor/Emisor',
                    'Archivo', 'CUIT Cliente', 'Fin CUIT', 'Cliente'
                ]
                data = data[columnas_necesarias]
                consolidado = pd.concat([consolidado , data], ignore_index=True)
            
    # Renombrar columnas
    consolidado.columns = [ 'Fecha' , 'Tipo' , 'Punto de Venta' , 'Número Desde' , 'Número Hasta' , 'Cód. Autorización' , 'Tipo Cambio' , 'Moneda' , 'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'Otros Tributos' , 'IVA' , 'Imp. Total' , 'Nro. Doc. Receptor/Emisor' , 'Denominación Receptor/Emisor' , 'Archivo' , 'CUIT Cliente' , 'Fin CUIT' , 'Cliente']

    #Eliminar las columas 'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA'
    consolidado.drop(['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA'], axis=1, inplace=True)

    #Cambiar de signo si es una Nota de Crédito (Tipo es numérico, no string)
    # Nota: El tipo 11 es factura C, tipo 13 es nota de crédito C, etc.
    # consolidado.loc[consolidado["Tipo"].str.contains("Nota de Crédito"), ['Imp. Total']] *= -1

    #Crear columna de 'MC' con los valores 'archivo' que van desde el caracter 5 al 8 en la Consolidado
    consolidado['MC'] = consolidado['Archivo'].str.split("-").str[1].str.strip()

    # Leer archivos JSON de RCEL
    for factura in archivos_PDF_JSON:
        #Si el existe el archivo, leerlo
        if os.path.isfile(factura):
            import json
            try:
                with open(factura, 'r', encoding='utf-8-sig') as f:
                    data_dict = json.load(f)
                    
                # Convertir a DataFrame (una sola fila)
                data_pdf = pd.DataFrame([data_dict])
                
                # si el datsaframe esta vacio, no hacer nada
                if len(data_pdf) > 0:
                    # Crear la columna 'Archivo PDF' con el ultimo elemento de 'factura' separado por "/"
                    data_pdf['Archivo PDF'] = factura.split("/")[-1]
                    
                    # Extraer información del nombre del archivo
                    # Formato: CUIT-COD-PtoVenta-NumFactura.json
                    partes = data_pdf["Archivo PDF"].str.split("-")
                    data_pdf['CUIT Cliente'] = partes.str[0].str.strip().astype(np.int64)
                    data_pdf['Fin CUIT'] = partes.str[0].str.strip().astype(np.int64)
                    
                    # Extraer Cliente del directorio padre
                    directorio_padre = factura.split("/")[-2]  # Ej: "20147130202_BUSTOS JOSE MARTIN"
                    cliente = directorio_padre.split("_", 1)[1] if "_" in directorio_padre else directorio_padre
                    data_pdf['Cliente'] = cliente
                    
                    Info_Facturas_PDF = pd.concat([Info_Facturas_PDF , data_pdf], ignore_index=True)
            except Exception as e:
                print(f"Error leyendo {factura}: {e}")
                continue

    # Construir AUX en consolidado para cruzar con los JSON de RCEL
    # Formato del JSON: CUIT_Emisor-COD(3)-PtoVenta(5)-Numero(8)
    # En MC: Fin CUIT es el CUIT del emisor (contribuyente)
    consolidado['AUX'] = (
        consolidado['Fin CUIT'].astype(int).astype(str) + "-" +
        consolidado['Tipo'].astype(int).astype(str).str.zfill(3) + "-" + 
        consolidado['Punto de Venta'].astype(int).astype(str).str.zfill(5) + "-" + 
        consolidado['Número Desde'].astype(int).astype(str).str.zfill(8)
    )

    # Merge con la tabla Info_Facturas_PDF 
    consolidado = pd.merge(consolidado , 
                           Info_Facturas_PDF[['AUX' , 'Desde' , 'Hasta' , 'Archivo PDF']] , 
                           how='left' , 
                           left_on='AUX' , 
                           right_on='AUX')

    # Crear la columna 'Cruzado' con valores 'Si' o 'No' dependiendo si se cruzó o no la información
    consolidado['Cruzado'] = ''
    consolidado.loc[consolidado['Archivo PDF'].notnull() , 'Cruzado'] = 'Si'
    consolidado.loc[consolidado['Archivo PDF'].isnull() , 'Cruzado'] = 'No'

    # Si las columnas 'Desde' y 'Hasta' son NaN entonces Eliminar todas filas donde la columna 'Fecha' no se encuentre entre el rango de fechas iniciales y finales
    ##########Consolidado = Consolidado[(Consolidado['Fecha'] >= fecha_inicial) & (Consolidado['Fecha'] <= fecha_final) & (Consolidado['Desde'].isnull()) & (Consolidado['Hasta'].isnull())]

    # Convertir la columna 'Fecha' a datetime primero
    # La Fecha del CSV viene en formato ISO (yyyy-mm-dd)
    consolidado['Fecha'] = pd.to_datetime(consolidado['Fecha'], format='ISO8601')
    
    # Convertir las columnas 'Desde' y 'Hasta' del JSON a datetime
    # Vienen en formato dd/mm/yyyy
    consolidado['Desde'] = pd.to_datetime(consolidado['Desde'], format='%d/%m/%Y', errors='coerce')
    consolidado['Hasta'] = pd.to_datetime(consolidado['Hasta'], format='%d/%m/%Y', errors='coerce')
    
    # Si las columnas 'Desde' y 'Hasta' son vacíos entonces toman el valor de 'Fecha'
    consolidado['Desde'] = consolidado['Desde'].fillna(consolidado['Fecha'])
    consolidado['Hasta'] = consolidado['Hasta'].fillna(consolidado['Fecha'])

    # Crear columna auxiliar en Consolidado con el máximo entre la variable fecha_inicial y la columna 'FECHA EMISION'
    consolidado['Fecha Inicial'] = fecha_inicial
    consolidado['Fecha_Inicial_max'] = consolidado[['Fecha Inicial' , 'Desde']].max(axis=1)
    del consolidado['Fecha Inicial']

    # Crear columna auxiliar en Consolidado con el mínimo entre la variable fecha_final y la columna 'HASTA'
    consolidado['Fecha Final'] = fecha_final
    consolidado['Fecha_Final_min'] = consolidado[['Fecha Final' , 'Hasta']].min(axis=1)
    del consolidado['Fecha Final']

    # Calcular los dias de diferencia entre 'FECHA EMISION' y 'HASTA'
    consolidado['Dias de facturación'] = consolidado['Hasta'] - consolidado['Desde'] 
    consolidado['Dias de facturación'] = consolidado['Dias de facturación'].dt.days +1

    # Crear una columa 'Días Efectivos' como la diferencia entre la columna 'Fecha_Final_min' y la columna 'Fecha_Inicial_max'
    consolidado['Días Efectivos'] = consolidado['Fecha_Final_min'] - consolidado['Fecha_Inicial_max']
    consolidado['Días Efectivos'] = consolidado['Días Efectivos'].dt.days +1

    # si los 'Días Efectivos' son menores a 0, entonces se reemplaza por 0
    consolidado.loc[consolidado['Días Efectivos'] < 0, 'Días Efectivos'] = 0

    # Crear una columna 'Importe por día' con el valor de la columna 'Imp. Total' dividido entre el valor de la columna 'Dias de facturación'
    consolidado['Importe por día'] = consolidado['Imp. Total'] / consolidado['Dias de facturación']

    # Multiplicar el valor de la columna 'Importe por día' por el valor de la columna 'Días Efectivos' y guardar el resultado en la columna 'Importe Prorrateado'
    consolidado['Importe Prorrateado'] = consolidado['Importe por día'] * consolidado['Días Efectivos']

    # Volver a mostrar las columnas 'Desde', 'Hasta' y 'Fecha' en formato fecha
    consolidado['Desde'] = consolidado['Desde'].dt.strftime('%d/%m/%Y')
    consolidado['Hasta'] = consolidado['Hasta'].dt.strftime('%d/%m/%Y')
    consolidado['Fecha'] = consolidado['Fecha'].dt.strftime('%d/%m/%Y')
    consolidado['Fecha_Inicial_max'] = consolidado['Fecha_Inicial_max'].dt.strftime('%d/%m/%Y')
    consolidado['Fecha_Final_min'] = consolidado['Fecha_Final_min'].dt.strftime('%d/%m/%Y')

    No_Cruzado = 0

    if 'No' in consolidado['Cruzado'].values:
        No_Cruzado = consolidado['Cruzado'].value_counts()['No']


    #Crear Tabla dinámica con los totales de las columnas  'Importe Prorrateado' por 'Archivo'
    TablaDinamica = pd.pivot_table(consolidado, values=['Importe Prorrateado' , 'Tipo'], index=['Cliente' , 'MC'], aggfunc={'Importe Prorrateado': 'sum' , 'Tipo': 'count'})


    # Renombrar la columna 'Tipo' por 'Cantidad de Comprobantes' de la TablaDinamica1 , TablaDinamica2 y TablaDinamica3
    TablaDinamica.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)

    # Buscar el valor de 'Importe Prorrateado' en la escala de categorias donde el valor esta en 'Ingresos brutos'
    TablaDinamica['Ingresos brutos máximos por la categoría'] = TablaDinamica['Importe Prorrateado'].apply(lambda x: categorias.loc[categorias['Ingresos brutos'] >= x, 'Ingresos brutos'].iloc[0])

    #Buscar la 'Categoría' en la escala de categorias donde el valor esta en 'Ingresos brutos máximos por la categoría'
    TablaDinamica['Categoría'] = TablaDinamica['Importe Prorrateado'].apply(lambda x: categorias.loc[categorias['Ingresos brutos'] >= x, 'Categoria'].iloc[0])

    # Exportar el Consolidado y la Tabla Dinámica a un archivo de Excel
    nombre_archivo = 'Reporte Recategorizaciones de Monotributistas.xlsx'
    Archivo_final = pd.ExcelWriter(nombre_archivo, engine='openpyxl')
    TablaDinamica.to_excel(Archivo_final, sheet_name='Tabla Dinámica', index=True)
    consolidado.to_excel(Archivo_final, sheet_name='Consolidado', index=False)
    Archivo_final.close()

    # Aplicar formatos del archivo
    wb = load_workbook(nombre_archivo)
    
    # Formatear hoja 'Tabla Dinámica'
    ws_tabla = wb['Tabla Dinámica']
    Aplicar_formato_encabezado(ws_tabla)
    Aplicar_formato_moneda(ws_tabla, 3, 3)
    Aplicar_formato_moneda(ws_tabla, 5, 5)
    Autoajustar_columnas(ws_tabla)
    Agregar_filtros(ws_tabla)
    
    # Formatear hoja 'Consolidado'
    ws_consolidado = wb['Consolidado']
    Aplicar_formato_encabezado(ws_consolidado)
    Aplicar_formato_moneda(ws_consolidado, 7, 10)
    Aplicar_formato_moneda(ws_consolidado, 27, 28)
    Alinear_columnas(ws_consolidado, 1, ws_consolidado.max_column, 'left')
    Autoajustar_columnas(ws_consolidado)
    Agregar_filtros(ws_consolidado)
    
    # Guardar el archivo con formato
    wb.save(nombre_archivo)
    wb.close()

    #Mostrar mensaje de finalización
    #showinfo(title="Finalizado", message=f"El archivo se ha generado correctamente.\n \nCantidad de Facturas no cruzados: {No_Cruzado}")



if __name__ == "__main__":
    import glob
    
    print("\n" + "="*80)
    print("INICIANDO PROCESO DE DESCARGA Y CONTROL DE MONOTRIBUTISTAS")
    print("="*80 + "\n")
    start = datetime.now()
    print(f"Inicio del proceso: {start.strftime('%Y-%m-%d %H:%M:%S')}")
    print("\n" + "="*80)
    
    archivo_base = "planilla-control-monotributistas.xlsx"
    df = pd.read_excel(archivo_base)

    # Mostrar el encabezado sin la columna sensible 'Clave_representante'
    print(df.head().drop(columns=['Clave_representante'], errors='ignore'))

    mrbot_user = os.getenv("MRBOT_USER")
    mrbot_api_key = os.getenv("MRBOT_API_KEY")
    base_url = os.getenv("BASE_URL")
    mis_comprobantes_endpoint = os.getenv("MIS_COMPROBANTES_ENDPOINT")
    rcel_endpoint = os.getenv("RCEL_ENDPOINT")
    downloads_mc_path = os.getenv("DOWNLOADS_MC_PATH", "descargas_mis_comprobantes")
    downloads_rcel_path = os.getenv("DOWNLOADS_RCEL_PATH", "descargas_rcel")

    # Procesar cada fila
    for index, row in df.iterrows():
        # Procesar Mis Comprobantes
        procesar_descarga_mc(row, mrbot_user, mrbot_api_key, base_url, mis_comprobantes_endpoint, downloads_mc_path)
        
        # Procesar RCEL
        procesar_descarga_rcel(row, mrbot_user, mrbot_api_key, base_url, rcel_endpoint, downloads_rcel_path)
    
    # Ejecutar control con los archivos descargados
    print("\n" + "="*80)
    print("INICIANDO CONTROL DE ARCHIVOS DESCARGADOS")
    print("="*80 + "\n")
    
    # Buscar archivos de Mis Comprobantes y RCEL
    archivos_mc = glob.glob(f"{downloads_mc_path}/**/extraido/*.csv", recursive=True)
    archivos_PDF = []  # No se usan archivos PDF directamente
    archivos_PDF_JSON = glob.glob(f"{downloads_rcel_path}/**/*.json", recursive=True)
    
    print(f"Archivos MC encontrados: {len(archivos_mc)}")
    print(f"Archivos JSON RCEL encontrados: {len(archivos_PDF_JSON)}")
    
    if archivos_mc or archivos_PDF_JSON:
        print("\nEjecutando función control...\n")
        control(archivos_mc, archivos_PDF, archivos_PDF_JSON)
        print("\n" + "="*80)
        print("CONTROL COMPLETADO")
        print("="*80)
        print("\nReporte generado: 'Reporte Recategorizaciones de Monotributistas.xlsx'")
    else:
        print("\nNo se encontraron archivos para procesar.")
        
    print("\n" + "="*80)
    print("PROCESO COMPLETADO")
    print("="*80 + "\n")
    end = datetime.now()
    print(f"Fin del proceso: {end.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Duración total: {end - start}")
    print("\n" + "="*80 + "\n")
        